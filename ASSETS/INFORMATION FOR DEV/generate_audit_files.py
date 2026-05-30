import os
import pandas as pd
import openpyxl

FOLDER = "/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/"
ORIGINAL_FILE = os.path.join(FOLDER, "MOHDrugPrice_17Nov2025_KSP_only_Composition_Indications (1) (1)(1).xlsx")
SAFE_FILE = os.path.join(FOLDER, "MOHDrugPrice_17Nov2025_KSP_enriched_SAFE.xlsx")

LOG_FILE = os.path.join(FOLDER, "enrichment_log_SAFE.xlsx")
REPORT_FILE = os.path.join(FOLDER, "validation_report_SAFE.txt")

def is_missing(val):
    if val is None: return True
    if pd.isna(val): return True
    v = str(val).strip().lower()
    if v in ["", "n/a", "na", "n.m", "n/m", "unknown", "not available", "null"]:
        return True
    return False

def main():
    print("Loading ORIGINAL and SAFE workbooks...")
    wb_orig = openpyxl.load_workbook(ORIGINAL_FILE, data_only=True)
    wb_safe = openpyxl.load_workbook(SAFE_FILE, data_only=True)
    
    warnings = []
    failed = False
    
    # Validation checks
    if len(wb_orig.sheetnames) != len(wb_safe.sheetnames):
        warnings.append("FAIL: Sheet count changed")
        failed = True
        
    for name1, name2 in zip(wb_orig.sheetnames, wb_safe.sheetnames):
        if name1 != name2:
            warnings.append(f"FAIL: Sheet name mismatch: {name1} vs {name2}")
            failed = True

    stats = {
        'total_scanned': 0,
        'total_filled': 0,
        'filled_cols': {},
        'invalid_changes': 0
    }
    
    logs = []
    
    for sheet_name in wb_orig.sheetnames:
        if sheet_name not in wb_safe:
            continue
            
        ws_orig = wb_orig[sheet_name]
        ws_safe = wb_safe[sheet_name]
        
        if ws_orig.max_row != ws_safe.max_row:
            warnings.append(f"FAIL: Row count changed in {sheet_name}: {ws_orig.max_row} vs {ws_safe.max_row}")
            failed = True
            
        if ws_orig.max_column != ws_safe.max_column:
            warnings.append(f"FAIL: Column count changed in {sheet_name}: {ws_orig.max_column} vs {ws_safe.max_column}")
            failed = True
            
        headers = [ws_orig.cell(row=1, column=c).value for c in range(1, ws_orig.max_column + 1)]
        headers_safe = [ws_safe.cell(row=1, column=c).value for c in range(1, ws_safe.max_column + 1)]
        
        if headers != headers_safe:
            warnings.append(f"FAIL: Headers changed in {sheet_name}")
            failed = True
            
        stats['total_scanned'] += max(0, ws_orig.max_row - 1)
        
        for r in range(2, ws_orig.max_row + 1):
            product_name = ws_orig.cell(row=r, column=1).value # Assuming Col A
            manufacturer = ws_orig.cell(row=r, column=2).value # Assuming Col B
            
            for c in range(1, ws_orig.max_column + 1):
                c_orig = ws_orig.cell(row=r, column=c)
                c_safe = ws_safe.cell(row=r, column=c)
                
                v_orig = c_orig.value
                v_safe = c_safe.value
                
                if v_orig != v_safe:
                    col_name = headers[c-1] if headers[c-1] else f"Column {c}"
                    
                    if is_missing(v_orig) and not is_missing(v_safe):
                        # Valid fill
                        stats['total_filled'] += 1
                        stats['filled_cols'][col_name] = stats['filled_cols'].get(col_name, 0) + 1
                        
                        logs.append({
                            'main_sheet': sheet_name,
                            'main_row_number': r,
                            'product_name': str(product_name),
                            'manufacturer_name': str(manufacturer),
                            'column_filled': str(col_name),
                            'old_value': str(v_orig) if v_orig is not None else "",
                            'new_value': str(v_safe),
                            'matched_source_file': 'not available from previous run',
                            'matched_source_sheet': 'not available from previous run',
                            'matched_source_row': 'not available from previous run',
                            'match_type': 'not available from previous run',
                            'confidence_score': 'not available from previous run',
                            'source_column_used': 'not available from previous run',
                            'notes': 'Verified through comparison audit'
                        })
                    else:
                        # Invalid change! Original was not missing or they both have values but are different
                        warnings.append(f"FAIL: Original non-empty value changed at R{r}C{c} ({col_name}): '{v_orig}' -> '{v_safe}'")
                        failed = True
                        stats['invalid_changes'] += 1
                        
                # Check formatting strictly if they weren't changed intentionally
                if c_orig.number_format != c_safe.number_format:
                    warnings.append(f"FAIL: Style/Format changed at R{r}C{c}: {c_orig.number_format} -> {c_safe.number_format}")
                    failed = True

    # Generate log
    df_logs = pd.DataFrame(logs)
    if not df_logs.empty:
        df_logs.to_excel(LOG_FILE, index=False)
    else:
        # Create empty log with correct columns
        df_logs = pd.DataFrame(columns=[
            'main_sheet', 'main_row_number', 'product_name', 'manufacturer_name', 
            'column_filled', 'old_value', 'new_value', 'matched_source_file', 
            'matched_source_sheet', 'matched_source_row', 'match_type', 
            'confidence_score', 'source_column_used', 'notes'
        ])
        df_logs.to_excel(LOG_FILE, index=False)

    # Generate report
    report = [
        "VALIDATION REPORT",
        "=================",
        f"Original file: {os.path.basename(ORIGINAL_FILE)}",
        f"Safe file: {os.path.basename(SAFE_FILE)}",
        "",
        f"Validation Result: {'FAIL' if failed else 'PASS'}",
        "",
        "Structural Checks:",
        f"- Same sheet names: {'No' if any('Sheet name mismatch' in w for w in warnings) else 'Yes'}",
        f"- Same dimensions: {'No' if any('count changed' in w for w in warnings) else 'Yes'}",
        f"- Same headers: {'No' if any('Headers changed' in w for w in warnings) else 'Yes'}",
        f"- No original non-empty values changed: {'No' if stats['invalid_changes'] > 0 else 'Yes'}",
        f"- No price cells changed: {'No' if any('Retail' in w or 'Sale' in w for w in warnings) else 'Yes'}",
        f"- No product/manufacturer cells changed: {'No' if any('PRODUCT' in w or 'MANUFACTURER' in w for w in warnings) else 'Yes'}",
        f"- No styles changed: {'No' if any('Style/Format' in w for w in warnings) else 'Yes'}",
        f"- Only missing cells were filled: {'No' if stats['invalid_changes'] > 0 else 'Yes'}",
        "",
        "Fill Statistics:",
        f"- Total rows scanned: {stats['total_scanned']}",
        f"- Total cells filled: {stats['total_filled']}",
        f"- Fill counts per column: {stats['filled_cols']}",
        "",
        "Warnings and Errors:"
    ]
    
    if not warnings:
        report.append("None. All checks passed.")
    else:
        # cap warnings
        report.extend(warnings[:50])
        if len(warnings) > 50:
            report.append(f"... and {len(warnings) - 50} more warnings.")
            
    with open(REPORT_FILE, 'w') as f:
        f.write("\n".join(report))
        
    print("Audit files generation complete.")
    print(f"Validation Result: {'FAIL' if failed else 'PASS'}")

if __name__ == "__main__":
    main()

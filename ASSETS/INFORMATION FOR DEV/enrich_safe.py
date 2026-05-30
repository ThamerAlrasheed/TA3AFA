import os
import re
import pandas as pd
import openpyxl
from copy import deepcopy

try:
    import rapidfuzz
    from rapidfuzz import fuzz
    def similarity(a, b):
        return fuzz.ratio(a, b)
except ImportError:
    from difflib import SequenceMatcher
    def similarity(a, b):
        return SequenceMatcher(None, a, b).ratio() * 100.0

# ----------------- Configuration -----------------
FOLDER = "/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/"
MAIN_FILE = os.path.join(FOLDER, "MOHDrugPrice_17Nov2025_KSP_only_Composition_Indications (1) (1)(1).xlsx")
OUT_FILE = os.path.join(FOLDER, "MOHDrugPrice_17Nov2025_KSP_enriched_SAFE.xlsx")
LOG_FILE = os.path.join(FOLDER, "enrichment_log_SAFE.xlsx")
REPORT_FILE = os.path.join(FOLDER, "validation_report_SAFE.txt")

# ----------------- Normalization -----------------
def normalize_manufacturer_name(name):
    if pd.isna(name) or name is None: return ""
    name = str(name).lower().strip()
    name = re.sub(r'\s+', ' ', name)
    # Remove legal suffixes
    suffixes = [r'\bltd\.?\b', r'\blimited\b', r'\bco\.?\b', r'\bcompany\b', 
                r'\bpharma\b', r'\bpharmaceuticals\b', r'\bindustries\b', 
                r'\binc\.?\b', r'\bllc\b', r'\bgmbh\b', r'\bag\b']
    for suf in suffixes:
        name = re.sub(suf, '', name)
    name = re.sub(r'[^\w\s]', '', name)
    return re.sub(r'\s+', ' ', name).strip()

def normalize_product_name(name, core=False):
    if pd.isna(name) or name is None: return ""
    name = str(name).lower().strip()
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'[^\w\s]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    
    if not core:
        return name
        
    # Core normalization: remove dosages and forms
    words = name.split()
    forms_stopwords = {'tablets', 'tablet', 'caps', 'capsules', 'capsule', 'syrup', 'vial', 
                       'injection', 'cream', 'ointment', 'drops', 'sachet', 'solution', 
                       'suspension', 'bottle', 'ampoule', 'powder', 'inhaler', 'suppositories',
                       'suppository', 'spray', 'lotion', 'gel', 'film', 'coated'}
                       
    core_words = []
    for w in words:
        if w in forms_stopwords: continue
        # ignore words with digits (e.g. 500mg, 10ml)
        if any(char.isdigit() for char in w): continue
        if w in ['mg', 'ml', 'g', 'mcg', 'iu', 'ui', 'mEq']: continue
        core_words.append(w)
        
    return " ".join(core_words).strip()

def is_missing(val):
    if val is None: return True
    if pd.isna(val): return True
    v = str(val).strip().lower()
    if v in ["", "n/a", "na", "n.m", "n/m", "unknown", "not available", "null"]:
        return True
    return False

# ----------------- Column Detection -----------------
def detect_source_columns(df_cols):
    mapping = {'pack_qty': None, 'ingredients': None, 'indications': None, 'interactions': None, 'product': None, 'manufacturer': None}
    cols = [str(c).lower().strip() for c in df_cols]
    
    # Exclude broad categories for Indications
    bad_ind_words = ['therapeutic area', 'primary care', 'consumer health', 'cardiovascular', 
                     'oncology', 'anti-infectives', 'pain', 'specialty care', 'category', 
                     'division', 'business unit', 'product type']
                     
    for i, c in enumerate(cols):
        if 'product name' in c or 'product / brand' in c or 'brand' == c or 'medicine / brand' in c or 'matched bayer product' in c:
            if not mapping['product']: mapping['product'] = df_cols[i]
            
        elif 'company' in c or 'manufacturer' in c:
            if not mapping['manufacturer']: mapping['manufacturer'] = df_cols[i]
            
        elif any(w in c for w in ['pack', 'presentation', 'dosage form', 'form', 'quantity', 'volume', 'tablet count', 'vial size']):
            if not mapping['pack_qty']: mapping['pack_qty'] = df_cols[i]
            
        elif any(w in c for w in ['composition', 'active ingredient', 'ingredient', 'generic name', 'substance', 'inn', 'api']):
            if not mapping['ingredients']: mapping['ingredients'] = df_cols[i]
            
        elif any(w in c for w in ['indication', 'uses', 'approved uses', 'used for']):
            # Ensure it's not a broad category
            if not any(bw in c for bw in bad_ind_words):
                if not mapping['indications']: mapping['indications'] = df_cols[i]
                
        elif 'interaction' in c:
            if not mapping['interactions']: mapping['interactions'] = df_cols[i]
            
    return mapping

# ----------------- Source Indexing -----------------
def build_source_index(folder):
    source_data = []
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and 'MOHDrugPrice' not in f and not f.startswith('~')]
    
    for f in files:
        f_path = os.path.join(folder, f)
        try:
            # try finding header
            df_test = pd.read_excel(f_path, nrows=5)
            header_row = 0
            for r in range(min(5, len(df_test))):
                row_vals = [str(v).lower() for v in df_test.iloc[r].values]
                if any('product' in v for v in row_vals):
                    header_row = r + 1
                    break
                    
            df = pd.read_excel(f_path, header=header_row if header_row > 0 else 0)
            mapping = detect_source_columns(df.columns)
            
            if not mapping['product']:
                # fallback if we missed product name
                for c in df.columns:
                    if 'product' in str(c).lower():
                        mapping['product'] = c
                        break
                        
            if not mapping['product']: continue
            
            file_mfg = f.split('_')[0].lower()
            if file_mfg == 'gsk': file_mfg = 'gsk'
            
            for idx, row in df.iterrows():
                p_val = row[mapping['product']]
                if pd.isna(p_val) or is_missing(p_val): continue
                
                m_val = row[mapping['manufacturer']] if mapping['manufacturer'] and not pd.isna(row[mapping['manufacturer']]) else file_mfg
                
                entry = {
                    'file': f,
                    'sheet': 'Sheet1', # simplified
                    'row': idx + (header_row if header_row > 0 else 1) + 2,
                    'product': str(p_val),
                    'norm_prod': normalize_product_name(p_val, core=False),
                    'core_prod': normalize_product_name(p_val, core=True),
                    'manufacturer': str(m_val) if not pd.isna(m_val) else file_mfg,
                    'norm_mfg': normalize_manufacturer_name(str(m_val) if not pd.isna(m_val) else file_mfg),
                    'pack_qty': str(row[mapping['pack_qty']]) if mapping['pack_qty'] and not pd.isna(row[mapping['pack_qty']]) else None,
                    'ingredients': str(row[mapping['ingredients']]) if mapping['ingredients'] and not pd.isna(row[mapping['ingredients']]) else None,
                    'indications': str(row[mapping['indications']]) if mapping['indications'] and not pd.isna(row[mapping['indications']]) else None,
                    'interactions': str(row[mapping['interactions']]) if mapping['interactions'] and not pd.isna(row[mapping['interactions']]) else None,
                    'used_cols': ", ".join([mapping[k] for k in ['pack_qty', 'ingredients', 'indications', 'interactions'] if mapping[k]])
                }
                source_data.append(entry)
        except Exception as e:
            print(f"Error reading {f}: {e}")
            
    return source_data

# ----------------- Matching Logic -----------------
def mfg_match_check(m_mfg, s_mfg):
    if not m_mfg or not s_mfg: return True
    if m_mfg in s_mfg or s_mfg in m_mfg: return True
    # if one of them is highly similar
    if similarity(m_mfg, s_mfg) >= 80: return True
    return False

def find_best_match(m_prod_orig, m_mfg_orig, sources):
    n_m_prod = normalize_product_name(m_prod_orig, core=False)
    c_m_prod = normalize_product_name(m_prod_orig, core=True)
    n_m_mfg = normalize_manufacturer_name(m_mfg_orig)
    
    exact_matches = []
    norm_matches = []
    fuzzy_matches = []
    
    for src in sources:
        is_mfg_match = mfg_match_check(n_m_mfg, src['norm_mfg'])
        
        # 1. Exact match
        if n_m_prod == src['norm_prod']:
            if is_mfg_match:
                exact_matches.append(src)
                continue
                
        # 2. Normalized match (Core match)
        if c_m_prod and c_m_prod == src['core_prod']:
            if is_mfg_match:
                norm_matches.append(src)
                continue
                
        # 3. Fuzzy match
        if is_mfg_match or not n_m_mfg: # require mfg match unless product name is unique, we enforce mfg match here to be conservative
            if is_mfg_match:
                score = similarity(n_m_prod, src['norm_prod'])
                if score >= 90:
                    fuzzy_matches.append((score, src))
                else:
                    # check core fuzzy
                    if c_m_prod and src['core_prod']:
                        c_score = similarity(c_m_prod, src['core_prod'])
                        if c_score >= 90:
                            fuzzy_matches.append((c_score, src))
                            
    # Decision logic
    if exact_matches:
        if len(exact_matches) == 1 or len(set(e['product'] for e in exact_matches)) == 1:
            return exact_matches[0], 'exact', 100
        return None, 'ambiguous', 0
        
    if norm_matches:
        if len(norm_matches) == 1 or len(set(e['product'] for e in norm_matches)) == 1:
            return norm_matches[0], 'normalized', 100
        return None, 'ambiguous', 0
        
    if fuzzy_matches:
        top_score = max(f[0] for f in fuzzy_matches)
        top_matches = [f[1] for f in fuzzy_matches if f[0] == top_score]
        if len(top_matches) == 1 or len(set(e['product'] for e in top_matches)) == 1:
            return top_matches[0], 'fuzzy', top_score
        return None, 'ambiguous', 0
        
    return None, 'skipped', 0

# ----------------- Validation -----------------
def validate_no_unwanted_changes(wb_orig, wb_mod):
    warnings = []
    failed = False
    
    if len(wb_orig.sheetnames) != len(wb_mod.sheetnames):
        warnings.append("Sheet count changed")
        failed = True
        
    for sheet_name in wb_orig.sheetnames:
        if sheet_name not in wb_mod:
            warnings.append(f"Sheet {sheet_name} missing")
            failed = True
            continue
            
        ws_orig = wb_orig[sheet_name]
        ws_mod = wb_mod[sheet_name]
        
        if ws_orig.max_row != ws_mod.max_row:
            warnings.append(f"Max rows changed in {sheet_name}: {ws_orig.max_row} vs {ws_mod.max_row}")
            failed = True
        if ws_orig.max_column != ws_mod.max_column:
            warnings.append(f"Max columns changed in {sheet_name}: {ws_orig.max_column} vs {ws_mod.max_column}")
            failed = True
            
        # Check all cells
        for r in range(1, ws_orig.max_row + 1):
            for c in range(1, ws_orig.max_column + 1):
                c_orig = ws_orig.cell(row=r, column=c)
                c_mod = ws_mod.cell(row=r, column=c)
                
                # if original had a value and we changed it
                if not is_missing(c_orig.value):
                    if c_orig.value != c_mod.value:
                        # Price columns are usually column D and E, Product is A, Mfg is B
                        warnings.append(f"Original non-missing cell changed at R{r}C{c}: '{c_orig.value}' -> '{c_mod.value}'")
                        failed = True
                        if len(warnings) > 10: 
                            warnings.append("... too many warnings, stopping")
                            return failed, warnings
                
                # Check formatting strictly (if possible in openpyxl)
                # openpyxl preserves it by default if we don't touch the cell
                if c_orig.number_format != c_mod.number_format:
                    warnings.append(f"Number format changed at R{r}C{c}: {c_orig.number_format} -> {c_mod.number_format}")
                    failed = True
                    if len(warnings) > 10: return failed, warnings

    return failed, warnings

# ----------------- Main Process -----------------
def main():
    print("Loading source files...")
    sources = build_source_index(FOLDER)
    print(f"Indexed {len(sources)} source products.")
    
    print("Loading ORIGINAL main file with openpyxl...")
    wb_orig = openpyxl.load_workbook(MAIN_FILE)
    wb_mod = openpyxl.load_workbook(MAIN_FILE)
    
    ws_mod = wb_mod.active # assume first sheet
    ws_orig = wb_orig.active
    
    headers = [ws_mod.cell(row=1, column=c).value for c in range(1, ws_mod.max_column + 1)]
    
    col_idx = {}
    for i, h in enumerate(headers):
        if not h: continue
        h_str = str(h).lower().strip()
        if 'product name' in h_str: col_idx['product'] = i + 1
        elif 'manufacturer name' in h_str: col_idx['manufacturer'] = i + 1
        elif 'pack qty' in h_str: col_idx['pack_qty'] = i + 1
        elif 'ingredients' in h_str or 'composition' in h_str: col_idx['ingredients'] = i + 1
        elif 'indications' in h_str: col_idx['indications'] = i + 1
        elif 'interactions' in h_str: col_idx['interactions'] = i + 1
        
    stats = {
        'total_scanned': ws_mod.max_row - 1,
        'total_matched': 0, 'exact_matches': 0, 'norm_matches': 0, 'fuzzy_matches': 0,
        'ambiguous': 0, 'skipped': 0, 'total_cells_filled': 0,
        'filled': {'pack_qty': 0, 'ingredients': 0, 'indications': 0, 'interactions': 0}
    }
    
    logs = []
    
    for r in range(2, ws_mod.max_row + 1):
        m_prod = ws_mod.cell(row=r, column=col_idx.get('product', 1)).value
        m_mfg = ws_mod.cell(row=r, column=col_idx.get('manufacturer', 2)).value
        
        if is_missing(m_prod):
            continue
            
        best_match, m_type, m_score = find_best_match(m_prod, m_mfg, sources)
        
        if m_type == 'ambiguous':
            stats['ambiguous'] += 1
            logs.append({'main_sheet': ws_mod.title, 'main_row_number': r, 'product_name_original': m_prod, 'manufacturer_original': m_mfg, 'skipped_reason': 'ambiguous', 'ambiguity_notes': 'Multiple possible matches found without clear winner'})
            continue
        elif not best_match:
            stats['skipped'] += 1
            logs.append({'main_sheet': ws_mod.title, 'main_row_number': r, 'product_name_original': m_prod, 'manufacturer_original': m_mfg, 'skipped_reason': 'no match found'})
            continue
            
        stats['total_matched'] += 1
        if m_type == 'exact': stats['exact_matches'] += 1
        elif m_type == 'normalized': stats['norm_matches'] += 1
        elif m_type == 'fuzzy': stats['fuzzy_matches'] += 1
        
        filled_now = []
        cols_used = []
        
        for k in ['pack_qty', 'ingredients', 'indications', 'interactions']:
            if k in col_idx and best_match[k]:
                c_idx = col_idx[k]
                old_val = ws_mod.cell(row=r, column=c_idx).value
                if is_missing(old_val):
                    ws_mod.cell(row=r, column=c_idx).value = best_match[k]
                    filled_now.append(headers[c_idx-1])
                    stats['filled'][k] += 1
                    stats['total_cells_filled'] += 1
                    
        logs.append({
            'main_sheet': ws_mod.title, 'main_row_number': r, 
            'product_name_original': m_prod, 'manufacturer_original': m_mfg,
            'matched_source_file': best_match['file'], 'matched_source_sheet': best_match['sheet'],
            'matched_source_row': best_match['row'], 'matched_product_name': best_match['product'],
            'matched_manufacturer': best_match['manufacturer'], 'match_type': m_type,
            'product_match_score': m_score, 'manufacturer_match_score': '',
            'columns_filled': ", ".join(filled_now), 'source_columns_used': best_match['used_cols']
        })
        
    print("Validating before saving...")
    failed, warnings = validate_no_unwanted_changes(wb_orig, wb_mod)
    
    report = [
        f"Original file: {os.path.basename(MAIN_FILE)}",
        f"Output file: {os.path.basename(OUT_FILE)}",
        f"Total rows scanned: {stats['total_scanned']}",
        f"Total rows matched: {stats['total_matched']}",
        f"Total cells filled: {stats['total_cells_filled']}",
        f"Fill counts per column: {stats['filled']}",
        f"Exact match count: {stats['exact_matches']}",
        f"Normalized match count: {stats['norm_matches']}",
        f"Fuzzy match count: {stats['fuzzy_matches']}",
        f"Skipped count: {stats['skipped']}",
        f"Ambiguous count: {stats['ambiguous']}",
        f"Validation result: {'FAIL' if failed else 'PASS'}",
        "Warnings:"
    ] + warnings
    
    with open(REPORT_FILE, 'w') as f:
        f.write("\n".join(report))
        
    # Write log
    pd.DataFrame(logs).to_excel(LOG_FILE, index=False)
    
    if not failed:
        wb_mod.save(OUT_FILE)
        print("SAFE ENRICHMENT COMPLETE")
    else:
        print("VALIDATION FAILED. Output not saved.")
        
    print(f"Total rows scanned: {stats['total_scanned']}")
    print(f"Total matched rows: {stats['total_matched']}")
    print(f"Total filled cells: {stats['total_cells_filled']}")
    print(f"Per-column fill counts: {stats['filled']}")
    print(f"Skipped/ambiguous count: {stats['skipped']} / {stats['ambiguous']}")
    print(f"Output files: {os.path.basename(OUT_FILE)}, {os.path.basename(LOG_FILE)}, {os.path.basename(REPORT_FILE)}")

if __name__ == "__main__":
    main()

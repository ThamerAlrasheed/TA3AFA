import openpyxl
import pandas as pd
import sys

# Paths
gsk_file = '/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/GSK_PDF_Product_Extraction.xlsx'
moh_file = '/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/MOHDrugPrice_17Nov2025_KSP_only_Composition_Indications.xlsx'

print('Loading GSK PDF extraction data...')
df_gsk = pd.read_excel(gsk_file, sheet_name='GSK PDF Extract')

# Let's organize the GSK data by its index or product name for easy lookup
# Clean product names to match our mapping
def clean_name(n):
    n = str(n).strip()
    if '[' in n:
        n = n.split('[')[0]
    if '.' in n:
        n = n.split('.')[0]
    return n.strip().upper()

df_gsk['CleanBrand'] = df_gsk['Product Name'].apply(clean_name)
gsk_data = {}
for idx, row in df_gsk.iterrows():
    brand = row['CleanBrand']
    gsk_data[brand] = {
        'ingredients': row['Ingredients (Composition)'],
        'indications': row['Indications']
    }

# Explicit mapped indices from pandas DataFrame (0-indexed, excluding header)
# We will map these pandas indices to the corresponding GSK brand name
pandas_index_mappings = {
    # Advair Diskus
    4911: 'ADVAIR DISKUS',
    4912: 'ADVAIR DISKUS',
    4913: 'ADVAIR DISKUS',
    # Advair HFA
    4914: 'ADVAIR HFA',
    4915: 'ADVAIR HFA',
    # Anoro Ellipta
    284: 'ANORO ELLIPTA',
    # Arexvy
    5592: 'AREXVY',
    # Breo Ellipta -> Relvar Ellipta
    4578: 'BREO ELLIPTA',
    4579: 'BREO ELLIPTA',
    # Engerix-B
    5598: 'ENGERIX-B. ENGERIX-B [HEPATITIS B VACCINE (REC...', # mapped clean name is 'ENGERIX-B'
    # Fluarix
    5599: 'FLUARIX',
    # Havrix
    5603: 'HAVRIX',
    5604: 'HAVRIX',
    # Hiberix
    5607: 'HIBERIX. HIBERIX [HAEMOPHILUS B CONJUGATE VACC...',
    # Imitrex -> Imigran
    2595: 'IMITREX',
    2596: 'IMITREX',
    2597: 'IMITREX',
    2598: 'IMITREX',
    # Lamictal (Using LAMICTAL ODT as primary, correcting composition to Lamotrigine)
    2923: 'LAMICTAL ODT',
    2925: 'LAMICTAL ODT',
    2927: 'LAMICTAL ODT',
    # Priorix
    5621: 'PRIORIX',
    5622: 'PRIORIX',
    # Rotarix
    5623: 'ROTARIX',
    5624: 'ROTARIX',
    5625: 'ROTARIX',
    # Serevent Diskus -> Serevent
    4916: 'SEREVENT DISKUS',
    4917: 'SEREVENT DISKUS',
    # Trelegy Ellipta
    5444: 'TRELEGY ELLIPTA',
    # Twinrix
    5630: 'TWINRIX [HEPATITIS A & HEPATITIS B (RECOMBINAN...',
    # Valtrex
    5655: 'VALTREX',
    5656: 'VALTREX',
    # Ventolin HFA -> Ventolin
    5720: 'VENTOLIN HFA',
    5721: 'VENTOLIN HFA',
    5722: 'VENTOLIN HFA',
    5724: 'VENTOLIN HFA',
    5725: 'VENTOLIN HFA',
    # Wellbutrin SR -> Wellbutrin XL
    5883: 'WELLBUTRIN SR',
    5884: 'WELLBUTRIN SR'
}

print('Loading MOH workbook...')
wb = openpyxl.load_workbook(moh_file)
sheet = wb['Sheet1']

# Let's determine column indices for:
# A: PRODUCT NAME
# F: Ingredients (Composition)
# G: Indications
# Let's inspect the headers in row 1
headers = [cell.value for cell in sheet[1]]
print('MOH Headers:', headers)

try:
    prod_col = headers.index('PRODUCT NAME') + 1
    ing_col = headers.index('Ingredients (Composition)') + 1
    ind_col = headers.index('Indications') + 1
except Exception as e:
    print('Error finding column headers:', e)
    sys.exit(1)

print(f'PRODUCT NAME column: {prod_col}')
print(f'Ingredients column: {ing_col}')
print(f'Indications column: {ind_col}')

# Let's update the rows
updated_count = 0
for p_idx, gsk_key in pandas_index_mappings.items():
    # Convert pandas index (0-indexed, excluding header) to openpyxl row (1-indexed, including header)
    row_num = p_idx + 2
    
    # Clean the brand key to match our clean names
    clean_key = clean_name(gsk_key)
    if clean_key not in gsk_data:
        print(f'WARNING: Could not find {clean_key} in GSK data!')
        continue
        
    gsk_item = gsk_data[clean_key]
    
    # Get cells
    prod_name = sheet.cell(row=row_num, column=prod_col).value
    ing_cell = sheet.cell(row=row_num, column=ing_col)
    ind_cell = sheet.cell(row=row_num, column=ind_col)
    
    # Prepare data
    ingredients = gsk_item['ingredients']
    indications = gsk_item['indications']
    
    # Correct Lamictal's ingredients from '12' to 'Lamotrigine'
    if 'LAMICTAL' in clean_key:
        ingredients = 'Lamotrigine is an antiepileptic drug (AED) of the phenyltriazine class.'
        
    # Write to cells
    ing_cell.value = ingredients
    ind_cell.value = indications
    
    print(f'Updated row {row_num}: \"{prod_name}\" -> mapped to GSK brand \"{clean_key}\"')
    updated_count += 1

print(f'Successfully updated {updated_count} rows in Sheet1!')

# Save the workbook
print('Saving MOH workbook...')
wb.save(moh_file)
print('Workbook saved successfully!')

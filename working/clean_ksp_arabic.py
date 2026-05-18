import openpyxl
import re
import sys

moh_file = '/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/MOHDrugPrice_17Nov2025_KSP_only_Composition_Indications.xlsx'

def clean_sentence(text):
    if not isinstance(text, str):
        return text
    
    # 1. Remove all Arabic characters (including presentation forms A and B)
    cleaned = re.sub(r'[\u0600-\u06FF\ufb50-\ufdff\ufe70-\ufeff]', '', text)
    
    # 2. Collapse multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    
    # 3. Clean up common parenthetical and punctuation glitches from mixture removal
    cleaned = re.sub(r'\(\s*\)', '', cleaned)
    cleaned = re.sub(r'\[\s*\]', '', cleaned)
    cleaned = re.sub(r'\(\s*,\s*', '(', cleaned)
    cleaned = re.sub(r'\s*,\s*\)', ')', cleaned)
    
    # 4. Clean double colons
    cleaned = re.sub(r':\s*:', ':', cleaned)
    cleaned = re.sub(r':\s*$', '', cleaned)
    cleaned = re.sub(r'^\s*:', '', cleaned)
    
    cleaned = cleaned.strip()
    
    # 5. Trim leading and trailing punctuation repeatedly
    while True:
        prev = cleaned
        cleaned = re.sub(r'^[.:,;\-\s\(\)]+', '', cleaned)
        cleaned = re.sub(r'[.:,;\-\s\(\)]+$', '', cleaned)
        if cleaned == prev:
            break
            
    # 6. Adjust spacing around punctuation
    cleaned = re.sub(r'\s+([:,;])', r'\1', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    # 7. Balance parentheses if necessary
    open_p = cleaned.count('(')
    close_p = cleaned.count(')')
    if open_p > close_p:
        cleaned += ')' * (open_p - close_p)
    elif close_p > open_p:
        cleaned = cleaned.rstrip(')')
        
    # 8. Ensure proper trailing dot if it ends with alphanumeric characters or closing parenthesis
    if cleaned and not cleaned.endswith('.'):
        if re.search(r'[a-zA-Z0-9)]$', cleaned):
            cleaned += '.'
            
    # 9. Clean up double dots at the end
    cleaned = re.sub(r'\.\.+$', '.', cleaned)
    
    return cleaned

print('Loading MOH workbook...')
wb = openpyxl.load_workbook(moh_file)
sheet = wb['Sheet1']

# Let's locate the column headers in row 1
headers = [cell.value for cell in sheet[1]]
print('Headers:', headers)

try:
    prod_col = headers.index('PRODUCT NAME') + 1
    mfr_col = headers.index('MANUFACTURER NAME') + 1
    ing_col = headers.index('Ingredients (Composition)') + 1
    ind_col = headers.index('Indications') + 1
except Exception as e:
    print('Error finding column headers:', e)
    sys.exit(1)

print(f'Columns -> Prod: {prod_col}, Mfr: {mfr_col}, Ing: {ing_col}, Ind: {ind_col}')

# Let's clean the KSP rows
ksp_cleaned = 0
for row_num in range(2, sheet.max_row + 1):
    mfr_val = sheet.cell(row=row_num, column=mfr_col).value
    
    # Check if the row belongs to KSP manufacturer
    if mfr_val and any(x in str(mfr_val).upper() for x in ['KSP', 'KUWAIT SAUDI']):
        prod_name = sheet.cell(row=row_num, column=prod_col).value
        
        # Get current cell values
        ing_cell = sheet.cell(row=row_num, column=ing_col)
        ind_cell = sheet.cell(row=row_num, column=ind_col)
        
        old_ing = ing_cell.value
        old_ind = ind_cell.value
        
        # We only clean if they are not 'N/M'
        if old_ing and old_ing != 'N/M':
            ing_cell.value = clean_sentence(str(old_ing))
            
        if old_ind and old_ind != 'N/M':
            ind_cell.value = clean_sentence(str(old_ind))
            
        if (old_ing and old_ing != ing_cell.value) or (old_ind and old_ind != ind_cell.value):
            print(f'Cleaned KSP row {row_num} (\"{prod_name}\"):')
            if old_ing != ing_cell.value:
                print(f'  Ingredients: {repr(old_ing)} -> {repr(ing_cell.value)}')
            if old_ind != ind_cell.value:
                print(f'  Indications: {repr(old_ind)} -> {repr(ind_cell.value)}')
            ksp_cleaned += 1

print(f'\nTotal KSP rows cleaned: {ksp_cleaned}')

print('Saving MOH workbook...')
wb.save(moh_file)
print('Workbook saved successfully!')

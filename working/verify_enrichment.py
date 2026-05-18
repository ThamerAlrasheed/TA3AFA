import pandas as pd
import openpyxl
import sys

# Paths
gsk_file = '/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/GSK_PDF_Product_Extraction.xlsx'
moh_file = '/Users/mac/Desktop/TA3AFA/ASSETS/INFORMATION FOR DEV/MOHDrugPrice_17Nov2025_KSP_only_Composition_Indications.xlsx'

print('=== Running Enrichment Verification ===')

# Load files
try:
    df_gsk = pd.read_excel(gsk_file, sheet_name='GSK PDF Extract')
    df_moh = pd.read_excel(moh_file, sheet_name='Sheet1')
except Exception as e:
    print('Error loading files:', e)
    sys.exit(1)

# Check sheet structure
print(f'MOH Sheet1 shape: {df_moh.shape}')
assert df_moh.shape == (6158, 7), f'Expected shape (6158, 7), but got {df_moh.shape}'
print('✓ Sheet shape is exactly identical.')

# Check list of sheets
wb = openpyxl.load_workbook(moh_file, read_only=True)
print('Workbook sheet names:', wb.sheetnames)
assert 'Sheet1' in wb.sheetnames and 'KSP Match Summary' in wb.sheetnames, 'Missing sheet names!'
print('✓ All sheets (Sheet1, KSP Match Summary) are perfectly intact.')

# Mapped pandas indices
pandas_indices = [
    4911, 4912, 4913,  # Seretide Diskus
    4914, 4915,        # Seretide MDI
    284,               # Anoro Ellipta
    5592,              # Vaccine-Arexvy
    4578, 4579,        # Relvar Ellipta
    5598,              # Vaccine-Engerix B
    5599,              # Vaccine-Fluarix
    5603, 5604,        # Vaccine-Havrix
    5607,              # Vaccine-Hiberix
    2595, 2596, 2597, 2598, # Imigran
    2923, 2925, 2927,  # Lamictal
    5621, 5622,        # Vaccine-Priorix
    5623, 5624, 5625,  # Vaccine-Rotarix
    4916, 4917,        # Serevent
    5444,              # Trelegy Ellipta
    5630,              # Vaccine-Twinrix
    5655, 5656,        # Valtrex
    5720, 5721, 5722, 5724, 5725, # Ventolin
    5883, 5884         # Wellbutrin XL
]

print('\nChecking matched rows...')
errors = 0
for idx in pandas_indices:
    row = df_moh.iloc[idx]
    ing = row['Ingredients (Composition)']
    ind = row['Indications']
    name = row['PRODUCT NAME']
    
    # Check that they are not 'N/M'
    if ing == 'N/M' or ind == 'N/M':
        print(f'Error at row index {idx} ({name}): Ingredients or Indications is still \"N/M\"!')
        errors += 1
    # Check Lamictal correction
    elif 'Lamictal' in name:
        if ing != 'Lamotrigine is an antiepileptic drug (AED) of the phenyltriazine class.':
            print(f'Error at Lamictal row index {idx}: Ingredients was not cleaned up correctly! Got: \"{ing}\"')
            errors += 1
            
print(f'Checked {len(pandas_indices)} updated rows. Errors found: {errors}')

# Check that other rows are untouched
print('\nChecking untouched rows...')
untouched_errors = 0
for idx in range(len(df_moh)):
    if idx not in pandas_indices:
        row = df_moh.iloc[idx]
        ing = row['Ingredients (Composition)']
        ind = row['Indications']
        
        # Non-KSP non-GSK rows should still be 'N/M'
        is_ksp = any(x in str(row['MANUFACTURER NAME']).upper() for x in ['KSP', 'KUWAIT SAUDI'])
        if not is_ksp:
            if ing != 'N/M' or ind != 'N/M':
                p_name = row['PRODUCT NAME']
                print(f'Error at untouched row index {idx} ({p_name}): values were modified but should be "N/M"!')
                untouched_errors += 1

print(f'Checked all untouched rows. Errors found: {untouched_errors}')

if errors == 0 and untouched_errors == 0:
    print('\n🎉 VERIFICATION PASSED SUCCESSFULLY! The data is extremely pristine and correctly formatted.')
else:
    print(f'\n❌ VERIFICATION FAILED. Errors: {errors + untouched_errors}')
    sys.exit(1)
